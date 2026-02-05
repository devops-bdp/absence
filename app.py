import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import plotly.express as px
import plotly.graph_objects as go
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4, landscape, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# Konfigurasi halaman
st.set_page_config(
    page_title="Audit & Analisis Absensi - Januari 2026",
    page_icon="📊",
    layout="wide"
)

# Judul aplikasi
st.title("📊 Audit & Analisis Data Absensi - Januari 2026")
st.markdown("---")

@st.cache_data
def load_data():
    """Load dan clean data dari CSV"""
    try:
        df = pd.read_csv('january.csv')
        
        # Filter baris yang bukan TOTAL (baris yang berisi "TOTAL FOR EMPLOYEE")
        df = df[~df['Employee ID'].astype(str).str.contains('TOTAL', na=False)]
        
        # Filter baris yang memiliki Employee ID valid (numeric)
        df = df[pd.to_numeric(df['Employee ID'], errors='coerce').notna()]
        
        # Convert Employee ID ke integer
        df['Employee ID'] = df['Employee ID'].astype(int)
        
        # Convert Date ke datetime
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
        
        # Parse waktu kerja (format HH:MM ke jam desimal)
        def parse_time_to_hours(time_str):
            """Convert waktu format HH:MM ke jam desimal"""
            if pd.isna(time_str) or time_str == '' or time_str == '00:00':
                return 0.0
            try:
                parts = str(time_str).split(':')
                if len(parts) == 2:
                    hours = int(parts[0])
                    minutes = int(parts[1])
                    return hours + (minutes / 60.0)
                return 0.0
            except:
                return 0.0
        
        # Parse Late In dan Early Out
        def parse_late_early(time_str):
            """Check apakah ada late in atau early out (bukan 00:00)"""
            if pd.isna(time_str) or time_str == '' or time_str == '00:00':
                return False
            try:
                parts = str(time_str).split(':')
                if len(parts) == 2:
                    hours = int(parts[0])
                    minutes = int(parts[1])
                    return hours > 0 or minutes > 0
                return False
            except:
                return False
        
        # Apply parsing
        df['Real Working Hour Decimal'] = df['Real Working Hour'].apply(parse_time_to_hours)
        df['Actual Working Hour Decimal'] = df['Actual Working Hour'].apply(parse_time_to_hours)
        df['Late In Decimal'] = df['Late In'].apply(parse_time_to_hours)
        df['Early Out Decimal'] = df['Early Out'].apply(parse_time_to_hours)
        df['Is Late In'] = df['Late In'].apply(parse_late_early)
        df['Is Early Out'] = df['Early Out'].apply(parse_late_early)
        
        # Tentukan apakah hadir (ada Check In atau Attendance Code = 'H')
        df['Is Present'] = (
            (df['Check In'].notna() & (df['Check In'] != '')) |
            (df['Attendance Code'] == 'H')
        )
        
        # Tentukan kategori status
        # Hari libur: Shift = 'dayoff'
        df['Is Dayoff'] = df['Shift'].str.contains('dayoff', case=False, na=False)
        
        # Cuti: Attendance Code = 'CT' atau Time Off Code = 'CT' atau Shift = 'Roster Leave'
        df['Is Leave'] = (
            (df['Attendance Code'] == 'CT') |
            (df['Time Off Code'] == 'CT') |
            (df['Shift'].str.contains('Roster Leave', case=False, na=False))
        )
        
        # Tidak hadir (absen): bukan hadir, bukan cuti, bukan hari libur
        df['Is Absent'] = (
            (~df['Is Present']) &
            (~df['Is Leave']) &
            (~df['Is Dayoff'])
        )
        
        return df
    except Exception as e:
        st.error(f"Error loading data: {str(e)}")
        return None

# Load data
df = load_data()

if df is not None:
    # Sidebar untuk filter
    st.sidebar.header("🔍 Filter Data")
    
    # Filter berdasarkan Branch - Default hanya HO Jakarta
    branches = sorted(df['Branch'].unique().tolist())
    default_branch = 'HO Jakarta' if 'HO Jakarta' in branches else branches[0] if branches else None
    selected_branch = st.sidebar.selectbox("Pilih Branch", branches, index=branches.index(default_branch) if default_branch and default_branch in branches else 0)
    
    # Filter berdasarkan Organization
    organizations = ['All'] + sorted(df['Organization'].unique().tolist())
    selected_org = st.sidebar.selectbox("Pilih Organization", organizations)
    
    # Fungsi untuk membuat report Excel
    def create_excel_report(employee_stats_df, checklist_df, filtered_df, branch, org):
        """Membuat report Excel lengkap dengan beberapa sheet"""
        output = BytesIO()
        writer = pd.ExcelWriter(output, engine='openpyxl')
        
        # Sheet 1: Ringkasan Statistik
        def format_hours_excel(hours):
            """Format jam desimal ke format yang mudah dibaca"""
            h = int(hours)
            m = int((hours - h) * 60)
            if m > 0:
                return f"{h:,} jam {m} menit"
            else:
                return f"{h:,} jam"
        
        total_emp = filtered_df['Employee ID'].nunique()
        total_pres = filtered_df['Is Present'].sum()
        total_abs = filtered_df['Is Absent'].sum()
        work_days = employee_stats_df['Work Days Bulan Ini'].iloc[0] if len(employee_stats_df) > 0 else 0
        total_work_day = total_emp * work_days
        attendance_pct = (total_pres / total_work_day * 100) if total_work_day > 0 else 0
        absent_pct = (total_abs / total_work_day * 100) if total_work_day > 0 else 0
        
        # Hitung total jam kerja dari employee_stats_df (Analisis Per Karyawan)
        total_jam_kerja_real = employee_stats_df['Total Jam Kerja (Real)'].sum() if 'Total Jam Kerja (Real)' in employee_stats_df.columns else filtered_df['Real Working Hour Decimal'].sum()
        # Total Jam Kerja Plant = Total Work Day × 8 jam
        total_jam_kerja_plant = total_work_day * 8
        total_jam_kerja_real_formatted = format_hours_excel(total_jam_kerja_real)
        total_jam_kerja_plant_formatted = format_hours_excel(total_jam_kerja_plant)
        
        summary_data = {
            'Metrik': [
                'Total Karyawan',
                'Work Day',
                'Total Work Day',
                'Total Kehadiran',
                'Total Kehadiran (%)',
                'Total Tidak Hadir',
                'Total Tidak Hadir (%)',
                'Total Jam Kerja',
                'Plant Jam Kerja'
            ],
            'Nilai': [
                total_emp,
                work_days,
                total_work_day,
                total_pres,
                f"{attendance_pct:.2f}%",
                total_abs,
                f"{absent_pct:.2f}%",
                total_jam_kerja_real_formatted,
                total_jam_kerja_plant_formatted
            ]
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Ringkasan Statistik', index=False)
        
        # Sheet 2: Analisis Per Karyawan (hapus kolom Branch dan Organization)
        employee_stats_df_export = employee_stats_df.drop(columns=['Branch', 'Organization'], errors='ignore')
        employee_stats_df_export.to_excel(writer, sheet_name='Analisis Per Karyawan', index=False)
        
        # Sheet 3: Checklist Compliance (hapus kolom Branch dan Organization)
        checklist_df_export = checklist_df.drop(columns=['Branch', 'Organization'], errors='ignore')
        checklist_df_export.to_excel(writer, sheet_name='Checklist Compliance', index=False)
        
        # Sheet 4: Detail Harian (sample untuk beberapa karyawan pertama) - hapus Branch dan Organization
        detail_cols = ['Date', 'Employee ID', 'Full Name', 'Job Position',
                      'Shift', 'Check In', 'Check Out', 'Late In', 'Early Out',
                      'Real Working Hour', 'Actual Working Hour', 'Attendance Code',
                      'Is Present', 'Is Absent', 'Is Dayoff', 'Is Leave']
        detail_df = filtered_df[detail_cols].copy()
        detail_df = detail_df.sort_values(['Full Name', 'Date'], ascending=[True, False])
        detail_df['Date'] = detail_df['Date'].dt.strftime('%Y-%m-%d')
        detail_df.to_excel(writer, sheet_name='Detail Harian', index=False)
        
        # Formatting Excel
        workbook = writer.book
        
        # Format Sheet Ringkasan Statistik
        ws1 = workbook['Ringkasan Statistik']
        ws1.column_dimensions['A'].width = 25
        ws1.column_dimensions['B'].width = 15
        
        # Header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Format Sheet Analisis Per Karyawan
        ws2 = workbook['Analisis Per Karyawan']
        for col in range(1, ws2.max_column + 1):
            ws2.column_dimensions[get_column_letter(col)].width = 18
        
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Format Sheet Checklist Compliance
        ws3 = workbook['Checklist Compliance']
        # Set width yang lebih spesifik untuk setiap kolom agar tidak tabrakan (tanpa Branch dan Organization)
        col_widths_checklist = {
            'Tanggal': 12,
            'ID': 10,
            'Nama': 25,
            'Posisi': 20,
            'Shift': 15,
            'Check In': 10,
            'Check Out': 10,
            'Jam Kerja (Format)': 15,
            'Jam Kerja (Desimal)': 12,
            '✅ Kerja 8 Jam/Hari': 18,
            '✅ Masuk 08:00 & Pulang 17:00': 25
        }
        
        # Set width untuk setiap kolom berdasarkan nama kolom
        for col_idx, col_name in enumerate(checklist_df_export.columns, start=1):
            col_letter = get_column_letter(col_idx)
            # Cari width yang sesuai dari dictionary, atau gunakan default 15
            width = col_widths_checklist.get(col_name, 15)
            ws3.column_dimensions[col_letter].width = width
        
        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Format Sheet Detail Harian
        ws4 = workbook['Detail Harian']
        for col in range(1, ws4.max_column + 1):
            ws4.column_dimensions[get_column_letter(col)].width = 15
        
        for cell in ws4[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Freeze first row untuk semua sheet
        for ws in workbook.worksheets:
            ws.freeze_panes = 'A2'
        
        writer.close()
        output.seek(0)
        return output
    
    # Fungsi untuk membuat report PDF
    def create_pdf_report(employee_stats_df, checklist_df, filtered_df, branch, org):
        """Membuat report PDF lengkap"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        story = []
        styles = getSampleStyleSheet()
        
        # Title Style
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1f4788'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Heading Style
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#1f4788'),
            spaceAfter=12,
            spaceBefore=12
        )
        
        # Normal Style
        normal_style = styles['Normal']
        
        # Title
        title = Paragraph("LAPORAN AUDIT & ANALISIS ABSENSI", title_style)
        story.append(title)
        
        # Info
        org_name = org if org != 'All' else 'Semua Organization'
        info_text = f"<b>Branch:</b> {branch} | <b>Organization:</b> {org_name}<br/>"
        info_text += f"<b>Periode:</b> {filtered_df['Date'].min().strftime('%d %B %Y')} - {filtered_df['Date'].max().strftime('%d %B %Y')}<br/>"
        info_text += f"<b>Tanggal Laporan:</b> {datetime.now().strftime('%d %B %Y %H:%M:%S')}"
        story.append(Paragraph(info_text, normal_style))
        story.append(Spacer(1, 0.3*inch))
        
        # Ringkasan Statistik
        story.append(Paragraph("1. RINGKASAN STATISTIK", heading_style))
        
        def format_hours_pdf(hours):
            """Format jam desimal ke format yang mudah dibaca"""
            h = int(hours)
            m = int((hours - h) * 60)
            if m > 0:
                return f"{h:,} jam {m} menit"
            else:
                return f"{h:,} jam"
        
        total_emp = filtered_df['Employee ID'].nunique()
        total_pres = filtered_df['Is Present'].sum()
        total_abs = filtered_df['Is Absent'].sum()
        work_days = employee_stats_df['Work Days Bulan Ini'].iloc[0] if len(employee_stats_df) > 0 else 0
        total_work_day = total_emp * work_days
        attendance_pct = (total_pres / total_work_day * 100) if total_work_day > 0 else 0
        absent_pct = (total_abs / total_work_day * 100) if total_work_day > 0 else 0
        
        # Hitung total jam kerja dari employee_stats_df (Analisis Per Karyawan)
        total_jam_kerja_real = employee_stats_df['Total Jam Kerja (Real)'].sum() if 'Total Jam Kerja (Real)' in employee_stats_df.columns else filtered_df['Real Working Hour Decimal'].sum()
        # Total Jam Kerja Plant = Total Work Day × 8 jam
        total_jam_kerja_plant = total_work_day * 8
        total_jam_kerja_real_formatted = format_hours_pdf(total_jam_kerja_real)
        total_jam_kerja_plant_formatted = format_hours_pdf(total_jam_kerja_plant)
        
        summary_data = [
            ['Metrik', 'Nilai'],
            ['Total Karyawan', str(total_emp)],
            ['Work Day', str(work_days)],
            ['Total Work Day', str(total_work_day)],
            ['Total Kehadiran', f"{total_pres} ({attendance_pct:.2f}%)"],
            ['Total Tidak Hadir', f"{total_abs} ({absent_pct:.2f}%)"],
            ['Total Jam Kerja', total_jam_kerja_real_formatted],
            ['Plant Jam Kerja', total_jam_kerja_plant_formatted]
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Analisis Per Karyawan (Top 20)
        story.append(Paragraph("2. ANALISIS PER KARYAWAN (Top 20)", heading_style))
        
        top_employees = employee_stats_df.head(20).copy()
        # Hapus kolom Branch dan Organization untuk PDF
        top_employees_export = top_employees.drop(columns=['Branch', 'Organization'], errors='ignore')
        emp_data = []
        emp_data.append(['ID', 'Nama', 'Hadir', 'Absen', 'Cuti', 'Late In', 'Early Out', 'Total Jam'])
        
        for _, row in top_employees.iterrows():
            emp_data.append([
                str(int(row['Employee ID'])),
                str(row['Full Name'])[:25] + '...' if len(str(row['Full Name'])) > 25 else str(row['Full Name']),
                str(int(row['Jumlah Hadir'])),
                str(int(row['Jumlah Absen'])),
                str(int(row['Jumlah Cuti'])),
                str(int(row['Jumlah Late In'])),
                str(int(row['Jumlah Early Out'])),
                str(row['Total Jam Kerja (Real) Formatted'])
            ])
        
        emp_table = Table(emp_data, colWidths=[0.6*inch, 2*inch, 0.5*inch, 0.5*inch, 0.5*inch, 0.6*inch, 0.6*inch, 0.8*inch])
        emp_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(emp_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Checklist Compliance Summary
        if len(checklist_df) > 0:
            story.append(Paragraph("3. RINGKASAN CHECKLIST COMPLIANCE", heading_style))
            
            # Hapus kolom Branch dan Organization untuk perhitungan (tidak mempengaruhi hasil)
            compliant_8jam = len(checklist_df[checklist_df['✅ Kerja 8 Jam/Hari'] == '✅'])
            compliant_8_17 = len(checklist_df[checklist_df['✅ Masuk 08:00 & Pulang 17:00'] == '✅'])
            compliant_both = len(checklist_df[
                (checklist_df['✅ Kerja 8 Jam/Hari'] == '✅') &
                (checklist_df['✅ Masuk 08:00 & Pulang 17:00'] == '✅')
            ])
            total_checklist = len(checklist_df)
            
            checklist_data = [
                ['Kriteria', 'Compliant', 'Tidak Compliant', 'Persentase'],
                ['Kerja 8 Jam/Hari', str(compliant_8jam), str(total_checklist - compliant_8jam), 
                 f"{(compliant_8jam/total_checklist*100) if total_checklist > 0 else 0:.1f}%"],
                ['Masuk 08:00 & Pulang 17:00', str(compliant_8_17), str(total_checklist - compliant_8_17),
                 f"{(compliant_8_17/total_checklist*100) if total_checklist > 0 else 0:.1f}%"],
                ['Keduanya Compliant', str(compliant_both), str(total_checklist - compliant_both),
                 f"{(compliant_both/total_checklist*100) if total_checklist > 0 else 0:.1f}%"]
            ]
            
            checklist_table = Table(checklist_data, colWidths=[2.5*inch, 1*inch, 1*inch, 1*inch])
            checklist_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
            ]))
            story.append(checklist_table)
            story.append(Spacer(1, 0.3*inch))
        
        # Footer
        story.append(Spacer(1, 0.2*inch))
        footer_text = f"<i>Laporan ini dibuat secara otomatis oleh sistem Audit & Analisis Absensi</i>"
        story.append(Paragraph(footer_text, normal_style))
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    # Fungsi helper untuk membuat PDF dari tabel
    def create_table_pdf(df, title, subtitle="", filename_prefix=""):
        """Membuat PDF dari DataFrame"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        story = []
        styles = getSampleStyleSheet()
        
        # Title Style
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1f4788'),
            spaceAfter=20,
            alignment=TA_CENTER
        )
        
        # Subtitle Style
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#666666'),
            spaceAfter=15,
            alignment=TA_CENTER
        )
        
        # Normal Style
        normal_style = styles['Normal']
        
        # Title
        story.append(Paragraph(title, title_style))
        if subtitle:
            story.append(Paragraph(subtitle, subtitle_style))
        
        # Info
        info_text = f"<b>Tanggal Laporan:</b> {datetime.now().strftime('%d %B %Y %H:%M:%S')}"
        story.append(Paragraph(info_text, normal_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Convert DataFrame to table
        if len(df) > 0:
            # Prepare data for table
            table_data = []
            
            # Header
            headers = [str(col) for col in df.columns]
            table_data.append(headers)
            
            # Data rows (limit to 100 rows untuk menghindari PDF terlalu besar)
            max_rows = 100
            df_display = df.head(max_rows) if len(df) > max_rows else df
            
            for _, row in df_display.iterrows():
                table_data.append([str(val) if pd.notna(val) else '' for val in row])
            
            # Create table
            table = Table(table_data)
            
            # Style table
            table.setStyle(TableStyle([
                # Header style
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                # Data style
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            
            story.append(table)
            
            # Note jika ada lebih dari max_rows
            if len(df) > max_rows:
                story.append(Spacer(1, 0.1*inch))
                note_text = f"<i>Catatan: Menampilkan {max_rows} dari {len(df)} baris. Untuk data lengkap, gunakan format Excel.</i>"
                story.append(Paragraph(note_text, normal_style))
        else:
            story.append(Paragraph("Tidak ada data untuk ditampilkan.", normal_style))
        
        # Footer
        story.append(Spacer(1, 0.2*inch))
        footer_text = f"<i>Laporan ini dibuat secara otomatis oleh sistem Audit & Analisis Absensi</i>"
        story.append(Paragraph(footer_text, normal_style))
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    # Fungsi helper untuk membuat PDF dari tabel
    def create_table_pdf(df, title, subtitle=""):
        """Membuat PDF dari DataFrame dalam format Landscape dengan layout yang rapi"""
        buffer = BytesIO()
        # Gunakan landscape orientation dengan margin yang lebih kecil
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=0.3*inch, bottomMargin=0.3*inch, 
                                leftMargin=0.2*inch, rightMargin=0.2*inch)
        story = []
        styles = getSampleStyleSheet()
        
        # Title Style
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=12,
            textColor=colors.HexColor('#1f4788'),
            spaceAfter=10,
            alignment=TA_CENTER
        )
        
        # Subtitle Style
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.HexColor('#666666'),
            spaceAfter=8,
            alignment=TA_CENTER
        )
        
        # Normal Style
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=7
        )
        
        # Title
        story.append(Paragraph(title, title_style))
        if subtitle:
            story.append(Paragraph(subtitle, subtitle_style))
        
        # Info
        info_text = f"<b>Tanggal Laporan:</b> {datetime.now().strftime('%d %B %Y %H:%M:%S')}"
        story.append(Paragraph(info_text, normal_style))
        story.append(Spacer(1, 0.1*inch))
        
        # Convert DataFrame to table
        if len(df) > 0:
            # Prepare data for table
            table_data = []
            
            # Header - wrap text untuk header yang panjang dan replace emoji
            headers = []
            for col in df.columns:
                col_str = str(col)
                # Replace emoji di header dengan teks yang jelas
                col_str = col_str.replace('✅', '').replace('❌', '').replace('✈️', '').replace('🏖️', '')
                col_str = col_str.strip()
                # Wrap header jika lebih dari 15 karakter
                if len(col_str) > 15:
                    # Split menjadi beberapa baris
                    words = col_str.split()
                    if len(words) > 1:
                        # Coba split di tengah
                        mid = len(words) // 2
                        header_text = ' '.join(words[:mid]) + '<br/>' + ' '.join(words[mid:])
                    else:
                        # Split karakter
                        mid = len(col_str) // 2
                        header_text = col_str[:mid] + '<br/>' + col_str[mid:]
                    headers.append(Paragraph(header_text, ParagraphStyle('Header', fontSize=7, alignment=TA_CENTER)))
                else:
                    headers.append(col_str)
            table_data.append(headers)
            
            # Data rows (limit to 200 rows untuk landscape)
            # Sort berdasarkan tanggal dari terendah ke tertinggi jika ada kolom tanggal
            max_rows = 200
            df_display = df.copy()
            
            # Cek apakah ada kolom tanggal untuk sorting
            date_cols = [col for col in df_display.columns if 'tanggal' in str(col).lower() or 'date' in str(col).lower()]
            if date_cols:
                try:
                    # Convert ke datetime jika belum
                    date_col = date_cols[0]
                    if not pd.api.types.is_datetime64_any_dtype(df_display[date_col]):
                        df_display[date_col] = pd.to_datetime(df_display[date_col], errors='coerce')
                    # Sort berdasarkan tanggal ascending (terendah ke tertinggi)
                    df_display = df_display.sort_values(by=date_col, ascending=True, na_position='last')
                except:
                    # Jika gagal, coba sort sebagai string
                    df_display = df_display.sort_values(by=date_cols[0], ascending=True, na_position='last')
            
            df_display = df_display.head(max_rows) if len(df_display) > max_rows else df_display
            
            # Track column indices yang perlu diwarnai (checklist columns)
            checklist_col_indices = []
            for i, col in enumerate(df.columns):
                col_str = str(col)
                # Deteksi kolom checklist berdasarkan nama kolom
                if any(keyword in col_str for keyword in ['Kerja 8 Jam', 'Masuk 08:00', 'Checklist', '✅', '❌']):
                    checklist_col_indices.append(i)
            
            for row_idx, row in df_display.iterrows():
                row_data = []
                for col_idx, val in enumerate(row):
                    if pd.notna(val):
                        col_name = df.columns[col_idx]
                        col_str = str(col_name).lower()
                        
                        # Format khusus untuk kolom tanggal - hapus bagian jam
                        if 'tanggal' in col_str or 'date' in col_str:
                            val_str = str(val)
                            # Jika masih ada format datetime dengan jam, ambil hanya bagian tanggal
                            if ' ' in val_str and ':' in val_str:
                                # Format: "2026-01-02 00:00:00" -> "2026-01-02"
                                val_str = val_str.split(' ')[0]
                            # Jika format datetime object, format ulang
                            try:
                                if pd.api.types.is_datetime64_any_dtype(type(val)) or isinstance(val, pd.Timestamp):
                                    val_str = pd.to_datetime(val).strftime('%Y-%m-%d')
                            except:
                                pass
                        else:
                            val_str = str(val)
                        
                        # Replace emoji dengan Y/N untuk PDF
                        val_str = val_str.replace('✅', 'Y').replace('❌', 'N')
                        val_str = val_str.replace('✈️', 'CUTI').replace('🏖️', 'LIBUR')
                        # Truncate cell value jika terlalu panjang (max 25 karakter)
                        if len(val_str) > 25:
                            row_data.append(val_str[:22] + '...')
                        else:
                            row_data.append(val_str)
                    else:
                        row_data.append('')
                table_data.append(row_data)
            
            # Calculate column widths berdasarkan konten
            num_cols = len(df.columns)
            page_width = landscape(A4)[0] - 0.4*inch  # Subtract margins
            available_width = page_width
            
            # Width khusus untuk kolom tertentu (untuk Checklist Compliance)
            special_widths = {
                'Tanggal': 1.0 * inch,
                'ID': 0.8 * inch,
                'Date': 1.0 * inch,
                'Employee ID': 0.8 * inch
            }
            
            # Hitung lebar optimal untuk setiap kolom berdasarkan konten
            col_widths = []
            for i, col in enumerate(df.columns):
                col_str = str(col)
                # Cek apakah ada width khusus untuk kolom ini
                if col_str in special_widths:
                    col_width = special_widths[col_str]
                else:
                    # Base width berdasarkan panjang header
                    header_len = len(col_str)
                    # Cek panjang maksimal data di kolom ini
                    max_data_len = df_display[col].astype(str).str.len().max() if len(df_display) > 0 else header_len
                    # Gunakan yang lebih panjang antara header dan data
                    content_len = max(header_len, max_data_len)
                    # Hitung lebar (minimum 0.6 inch, maksimum 2 inch per kolom)
                    col_width = min(max(content_len * 0.08, 0.6), 2.0) * inch
                col_widths.append(col_width)
            
            # Normalisasi lebar kolom agar total sesuai dengan page width
            total_width = sum(col_widths)
            if total_width > available_width:
                # Scale down semua kolom
                scale_factor = available_width / total_width
                col_widths = [w * scale_factor for w in col_widths]
            elif total_width < available_width:
                # Distribute extra space
                extra = (available_width - total_width) / num_cols
                col_widths = [w + extra for w in col_widths]
            
            # Create table dengan repeatRows untuk header
            table = Table(table_data, colWidths=col_widths, repeatRows=1)
            
            # Style table yang lebih rapi
            table_style = TableStyle([
                # Header style
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 7),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('TOPPADDING', (0, 0), (-1, 0), 6),
                # Data style default
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 6.5),
                ('GRID', (0, 0), (-1, -1), 0.3, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
                ('TOPPADDING', (0, 1), (-1, -1), 3),
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ])
            
            # Tambahkan warna untuk kolom checklist (hijau untuk Y, merah untuk N)
            # Hapus ROWBACKGROUNDS untuk kolom checklist agar warna tidak tertimpa
            if checklist_col_indices:
                for col_idx in checklist_col_indices:
                    # Override ROWBACKGROUNDS untuk kolom checklist
                    for row_idx in range(1, len(table_data)):
                        # Set background default dulu (untuk cell yang bukan Y atau N)
                        table_style.add('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), colors.beige)
            
            # Loop melalui data untuk mewarnai cell berdasarkan nilai Y/N
            for col_idx in checklist_col_indices:
                for row_idx in range(1, len(table_data)):
                    cell_value = str(table_data[row_idx][col_idx]).strip().upper()
                    if cell_value == 'Y':
                        # Hijau untuk Y
                        table_style.add('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), colors.HexColor('#90EE90'))  # Light green
                        table_style.add('TEXTCOLOR', (col_idx, row_idx), (col_idx, row_idx), colors.HexColor('#006400'))  # Dark green
                        table_style.add('FONTNAME', (col_idx, row_idx), (col_idx, row_idx), 'Helvetica-Bold')
                    elif cell_value == 'N':
                        # Merah untuk N
                        table_style.add('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), colors.HexColor('#FFB6C1'))  # Light pink/coral
                        table_style.add('TEXTCOLOR', (col_idx, row_idx), (col_idx, row_idx), colors.HexColor('#8B0000'))  # Dark red
                        table_style.add('FONTNAME', (col_idx, row_idx), (col_idx, row_idx), 'Helvetica-Bold')
            
            table.setStyle(table_style)
            
            story.append(table)
            
            # Note jika ada lebih dari max_rows
            if len(df) > max_rows:
                story.append(Spacer(1, 0.08*inch))
                note_text = f"<i>Catatan: Menampilkan {max_rows} dari {len(df)} baris. Untuk data lengkap, gunakan format Excel.</i>"
                story.append(Paragraph(note_text, normal_style))
        else:
            story.append(Paragraph("Tidak ada data untuk ditampilkan.", normal_style))
        
        # Footer
        story.append(Spacer(1, 0.1*inch))
        footer_text = f"<i>Laporan ini dibuat secara otomatis oleh sistem Audit & Analisis Absensi</i>"
        story.append(Paragraph(footer_text, normal_style))
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    # Fungsi untuk membuat PDF lengkap Detail Per Karyawan
    def create_employee_full_pdf(employee_info_df, attendance_stats_df, detail_display_df, employee_name, emp_data, date_start, date_end):
        """Membuat PDF lengkap untuk Detail Per Karyawan"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=0.3*inch, bottomMargin=0.3*inch, 
                                leftMargin=0.2*inch, rightMargin=0.2*inch)
        story = []
        styles = getSampleStyleSheet()
        
        # Title Style
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=14,
            textColor=colors.HexColor('#1f4788'),
            spaceAfter=10,
            alignment=TA_CENTER
        )
        
        # Heading Style
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=11,
            textColor=colors.HexColor('#1f4788'),
            spaceAfter=8,
            spaceBefore=8
        )
        
        # Normal Style
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=8
        )
        
        # Title
        story.append(Paragraph(f"DETAIL PER KARYAWAN - {employee_name}", title_style))
        subtitle_text = f"Periode: {date_start} - {date_end} | Tanggal Laporan: {datetime.now().strftime('%d %B %Y %H:%M:%S')}"
        story.append(Paragraph(subtitle_text, normal_style))
        story.append(Spacer(1, 0.15*inch))
        
        # 1. Informasi Karyawan
        story.append(Paragraph("1. INFORMASI KARYAWAN", heading_style))
        info_data = [['Informasi', 'Nilai']]
        for _, row in employee_info_df.iterrows():
            info_data.append([str(row['Informasi']), str(row['Nilai'])])
        
        info_table = Table(info_data, colWidths=[2.5*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTSIZE', (0, 1), (-1, -1), 7.5),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 0.2*inch))
        
        # 2. Statistik Absensi
        story.append(Paragraph("2. STATISTIK ABSENSI", heading_style))
        stats_data = [['Metrik', 'Nilai']]
        for _, row in attendance_stats_df.iterrows():
            stats_data.append([str(row['Metrik']), str(row['Nilai'])])
        
        stats_table = Table(stats_data, colWidths=[3*inch, 3.5*inch])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTSIZE', (0, 1), (-1, -1), 7.5),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
        ]))
        story.append(stats_table)
        story.append(Spacer(1, 0.2*inch))
        
        # 3. Detail Harian
        story.append(Paragraph("3. DETAIL HARIAN", heading_style))
        if len(detail_display_df) > 0:
            # Hapus kolom Branch dan Organization jika ada
            detail_display_df_export = detail_display_df.drop(columns=['Branch', 'Organization'], errors='ignore')
            # Prepare data untuk tabel detail harian
            detail_data = []
            headers = [str(col) for col in detail_display_df_export.columns]
            detail_data.append(headers)
            
            # Sort berdasarkan tanggal ascending
            detail_sorted = detail_display_df_export.copy()
            if 'Tanggal' in detail_sorted.columns:
                try:
                    detail_sorted['Tanggal_Sort'] = pd.to_datetime(detail_sorted['Tanggal'].str[:10], errors='coerce')
                    detail_sorted = detail_sorted.sort_values('Tanggal_Sort', ascending=True, na_position='last')
                    detail_sorted = detail_sorted.drop('Tanggal_Sort', axis=1)
                except:
                    detail_sorted = detail_sorted.sort_values('Tanggal', ascending=True)
            
            for _, row in detail_sorted.iterrows():
                row_data = []
                for val in row:
                    if pd.notna(val):
                        val_str = str(val)
                        val_str = val_str.replace('✅', 'Y').replace('❌', 'N')
                        val_str = val_str.replace('✈️', 'CUTI').replace('🏖️', 'LIBUR')
                        if len(val_str) > 25:
                            row_data.append(val_str[:22] + '...')
                        else:
                            row_data.append(val_str)
                    else:
                        row_data.append('')
                detail_data.append(row_data)
            
            # Calculate column widths
            num_cols = len(detail_display_df_export.columns)
            page_width = landscape(A4)[0] - 0.4*inch
            col_width = page_width / num_cols if num_cols > 0 else 1*inch
            col_widths = [max(col_width, 0.7*inch) for _ in range(num_cols)]
            
            detail_table = Table(detail_data, colWidths=col_widths, repeatRows=1)
            detail_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 7),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 6.5),
                ('GRID', (0, 0), (-1, -1), 0.3, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ]))
            story.append(detail_table)
        else:
            story.append(Paragraph("Tidak ada data detail harian.", normal_style))
        
        # Footer
        story.append(Spacer(1, 0.15*inch))
        footer_text = f"<i>Laporan ini dibuat secara otomatis oleh sistem Audit & Analisis Absensi</i>"
        story.append(Paragraph(footer_text, normal_style))
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    # Filter data - Hanya HO Jakarta
    filtered_df = df[df['Branch'] == selected_branch].copy()
    if selected_org != 'All':
        filtered_df = filtered_df[filtered_df['Organization'] == selected_org]
    
    # Hitung work days bulan ini (hari kerja Senin-Jumat) - dipindahkan ke atas untuk digunakan di Ringkasan Statistik
    def calculate_work_days(year, month):
        """Hitung jumlah hari kerja (Senin-Jumat) dalam bulan tertentu"""
        from calendar import monthrange
        import datetime
        
        # Dapatkan jumlah hari dalam bulan
        num_days = monthrange(year, month)[1]
        
        # Hitung hari kerja (Senin=0, Jumat=4)
        work_days = 0
        for day in range(1, num_days + 1):
            date = datetime.date(year, month, day)
            # 0 = Senin, 4 = Jumat
            if date.weekday() < 5:  # Senin sampai Jumat
                work_days += 1
        
        return work_days
    
    # Dapatkan tahun dan bulan dari data untuk work days
    if not filtered_df['Date'].empty:
        first_date = filtered_df['Date'].min()
        year = first_date.year
        month = first_date.month
        work_days_month = calculate_work_days(year, month)
    else:
        work_days_month = 0
    
    # Ringkasan statistik
    st.header("📈 Ringkasan Statistik")
    
    # Fungsi untuk format jam
    def format_hours(hours):
        """Format jam desimal ke format yang mudah dibaca"""
        h = int(hours)
        m = int((hours - h) * 60)
        # Format dengan pemisah ribuan dan menit terpisah
        if m > 0:
            return f"{h:,} jam {m} menit"
        else:
            return f"{h:,} jam"
    
    def format_hours_simple(hours):
        """Format jam desimal ke format sederhana (untuk help text)"""
        return f"{hours:,.2f} jam"
    
    # Hitung semua metrik
    total_employees = filtered_df['Employee ID'].nunique()
    total_records = len(filtered_df)
    total_present = filtered_df['Is Present'].sum()
    total_absent = filtered_df['Is Absent'].sum()
    total_late = filtered_df['Is Late In'].sum()
    total_early_out = filtered_df['Is Early Out'].sum()
    total_work_day = total_employees * work_days_month
    attendance_percentage = (total_present / total_work_day * 100) if total_work_day > 0 else 0
    absent_percentage = (total_absent / total_work_day * 100) if total_work_day > 0 else 0
    
    # Hitung employee_stats terlebih dahulu untuk mendapatkan total jam kerja dari Analisis Per Karyawan
    employee_stats_temp = filtered_df.groupby(['Employee ID', 'Full Name', 'Branch', 'Organization', 'Job Position']).agg({
        'Real Working Hour Decimal': 'sum'
    }).reset_index()
    
    # Hitung total jam kerja dari employee_stats (Analisis Per Karyawan)
    total_jam_kerja_real = employee_stats_temp['Real Working Hour Decimal'].sum()
    # Total Jam Kerja Plant = Total Work Day × 8 jam (sama seperti di Analisis Per Karyawan)
    total_jam_kerja_plant = total_work_day * 8
    total_jam_kerja_real_formatted = format_hours(total_jam_kerja_real)
    total_jam_kerja_plant_formatted = format_hours(total_jam_kerja_plant)
    
    # Tampilkan metrik dalam 2 baris
    col_stat1, col_stat2, col_stat3, col_stat4, col_stat5 = st.columns(5)
    
    with col_stat1:
        st.metric("Total Karyawan", total_employees, help="Formula: COUNT(DISTINCT Employee ID)")
    with col_stat2:
        st.metric("Work Day", work_days_month, help="Formula: Jumlah hari kerja (Senin-Jumat) dalam bulan")
    with col_stat3:
        st.metric("Total Work Day", total_work_day, help="Formula: Total Karyawan × Work Day")
    with col_stat4:
        st.metric("Total Kehadiran", f"{total_present} ({attendance_percentage:.1f}%)", help=f"Formula: SUM(Is Present)\nPersentase: (Total Kehadiran / Total Work Day) × 100%")
    with col_stat5:
        st.metric("Total Tidak Hadir", f"{total_absent} ({absent_percentage:.1f}%)", help=f"Formula: SUM(Is Absent)\nPersentase: (Total Tidak Hadir / Total Work Day) × 100%")
    
    # Baris kedua untuk Total Jam Kerja
    col_stat6, col_stat7 = st.columns(2)
    
    # Hitung jumlah record yang digunakan untuk perhitungan
    total_records_with_hours = len(filtered_df[filtered_df['Real Working Hour Decimal'] > 0])
    
    with col_stat6:
        help_text_real = (
            f"Formula: SUM(Real Working Hour Decimal)\n"
            f"Penjelasan: Menjumlahkan semua jam kerja real dari semua record absensi\n"
            f"Total Record: {len(filtered_df):,} record\n"
            f"Record dengan Jam Kerja: {total_records_with_hours:,} record\n"
            f"Hasil {total_jam_kerja_real_formatted} menunjukkan total jam kerja aktual dari semua record"
        )
        st.metric("Total Jam Kerja", total_jam_kerja_real_formatted, help=help_text_real)
    
    with col_stat7:
        help_text_plant = (
            f"Formula: Total Work Day × 8 jam\n"
            f"Penjelasan: Total jam kerja ideal (Plant) = {total_work_day} × 8 = {total_jam_kerja_plant:.2f} jam\n"
            f"Ini adalah total jam kerja ideal jika semua karyawan bekerja 8 jam per hari kerja"
        )
        st.metric("Plant Jam Kerja", total_jam_kerja_plant_formatted, help=help_text_plant)
    
    # Perbandingan Plant vs Actual
    st.markdown("### 📊 Perbandingan Plant vs Actual")
    
    # Hitung Plant (ideal): Total Work Day × 8 jam
    # Total Work Day sudah = Total Karyawan × Work Day
    # Jadi Plant = Total Work Day × 8 jam (tidak perlu dikalikan lagi dengan Total Karyawan)
    plant_total_hours = total_work_day * 8
    plant_total_formatted = format_hours(plant_total_hours)
    
    # Hitung selisih dan persentase
    selisih_hours = total_jam_kerja_real - plant_total_hours
    persentase = (total_jam_kerja_real / plant_total_hours * 100) if plant_total_hours > 0 else 0
    selisih_formatted = format_hours(abs(selisih_hours))
    delta_color = "normal" if selisih_hours >= 0 else "inverse"
    
    col_plant1, col_plant2, col_plant3 = st.columns(3)
    
    with col_plant1:
        st.metric(
            "Plant (Ideal)",
            plant_total_formatted,
            help=f"Formula: Total Work Day × 8 jam\n= {total_work_day} × 8 = {plant_total_hours:.2f} jam\nTotal Work Day = {total_employees} karyawan × {work_days_month} hari = {total_work_day}\nIni adalah total jam kerja ideal jika semua karyawan bekerja 8 jam per hari kerja"
        )
    
    with col_plant2:
        st.metric(
            "Actual (Real)",
            total_jam_kerja_real_formatted,
            help=f"Formula: SUM dari Analisis Per Karyawan (Real Working Hour Decimal)\n= {total_jam_kerja_real:.2f} jam\nIni adalah total jam kerja aktual dari data absensi yang dihitung per karyawan"
        )
    
    with col_plant3:
        st.metric(
            "Selisih",
            f"{selisih_formatted} ({persentase:.1f}%)",
            delta=f"{selisih_hours:+.2f} jam",
            delta_color=delta_color,
            help=f"Selisih: Actual - Plant\n= {total_jam_kerja_real:.2f} - {plant_total_hours:.2f} = {selisih_hours:+.2f} jam\nPersentase: (Actual / Plant) × 100% = {persentase:.1f}%"
        )
    
    # Download Ringkasan Statistik
    summary_data = {
        'Metrik': ['Total Karyawan', 'Work Day', 'Total Work Day', 'Total Kehadiran', 'Total Kehadiran (%)', 'Total Tidak Hadir', 'Total Tidak Hadir (%)', 'Total Jam Kerja', 'Plant Jam Kerja'],
        'Nilai': [total_employees, work_days_month, total_work_day, total_present, f"{attendance_percentage:.2f}%", total_absent, f"{absent_percentage:.2f}%", total_jam_kerja_real_formatted, total_jam_kerja_plant_formatted]
    }
    summary_df = pd.DataFrame(summary_data)
    csv_summary = summary_df.to_csv(index=False)
    
    col_summary1, col_summary2 = st.columns(2)
    with col_summary1:
        st.download_button(
            label="📥 Download Ringkasan Statistik (CSV)",
            data=csv_summary,
            file_name=f"ringkasan_statistik_{selected_branch}_{selected_org}.csv",
            mime="text/csv",
            key='download_summary_csv'
        )
    with col_summary2:
        pdf_summary = create_table_pdf(
            summary_df,
            "RINGKASAN STATISTIK",
            f"Branch: {selected_branch} | Organization: {selected_org if selected_org != 'All' else 'Semua'}"
        )
        st.download_button(
            label="📄 Download Ringkasan Statistik (PDF)",
            data=pdf_summary.getvalue(),
            file_name=f"ringkasan_statistik_{selected_branch}_{selected_org}.pdf",
            mime="application/pdf",
            key='download_summary_pdf'
        )
    
    st.markdown("---")
    
    # Analisis per karyawan
    st.header("👥 Analisis Per Karyawan")
    
    # work_days_month sudah dihitung di atas untuk Ringkasan Statistik
    
    # Group by Employee
    employee_stats = filtered_df.groupby(['Employee ID', 'Full Name', 'Branch', 'Organization', 'Job Position']).agg({
        'Is Present': 'sum',
        'Is Absent': 'sum',
        'Is Dayoff': 'sum',
        'Is Leave': 'sum',
        'Is Late In': 'sum',
        'Is Early Out': 'sum',
        'Real Working Hour Decimal': 'sum',
        'Late In Decimal': 'sum',
        'Early Out Decimal': 'sum'
    }).reset_index()
    
    employee_stats.columns = [
        'Employee ID', 'Full Name', 'Branch', 'Organization', 'Job Position',
        'Jumlah Hadir', 'Jumlah Absen', 'Jumlah Hari Libur', 'Jumlah Cuti',
        'Jumlah Late In', 'Jumlah Early Out',
        'Total Jam Kerja (Real)',
        'Total Jam Late In', 'Total Jam Early Out'
    ]
    
    # Tambahkan work days bulan ini
    employee_stats['Work Days Bulan Ini'] = work_days_month
    
    # Hitung Total Jam Kerja (Plant) = Work Days Bulan Ini × 8 jam
    employee_stats['Total Jam Kerja (Plant)'] = employee_stats['Work Days Bulan Ini'] * 8
    
    # Format jam kerja
    def format_hours(hours):
        """Format jam desimal ke format yang mudah dibaca"""
        h = int(hours)
        m = int((hours - h) * 60)
        if m > 0:
            return f"{h:,} jam {m} menit"
        else:
            return f"{h:,} jam"
    
    employee_stats['Total Jam Kerja (Real) Formatted'] = employee_stats['Total Jam Kerja (Real)'].apply(format_hours)
    employee_stats['Total Jam Kerja (Plant) Formatted'] = employee_stats['Total Jam Kerja (Plant)'].apply(format_hours)
    employee_stats['Total Jam Late In Formatted'] = employee_stats['Total Jam Late In'].apply(format_hours)
    employee_stats['Total Jam Early Out Formatted'] = employee_stats['Total Jam Early Out'].apply(format_hours)
    
    # Hitung checklist dan kekurangan jam kerja
    # Checklist: ✅ (hijau) jika Real >= Plant, ❌ (merah) jika Real < Plant
    employee_stats['Checklist Plant'] = employee_stats.apply(
        lambda row: '✅' if row['Total Jam Kerja (Real)'] >= row['Total Jam Kerja (Plant)'] else '❌',
        axis=1
    )
    
    # Kekurangan jam kerja: Plant - Real (jika Real < Plant), 0 jika sudah memenuhi
    employee_stats['Kekurangan Jam Kerja'] = employee_stats.apply(
        lambda row: max(0, row['Total Jam Kerja (Plant)'] - row['Total Jam Kerja (Real)']),
        axis=1
    )
    
    # Format kekurangan jam kerja
    employee_stats['Kekurangan Jam Kerja Formatted'] = employee_stats['Kekurangan Jam Kerja'].apply(
        lambda x: format_hours(x) if x > 0 else '0 jam'
    )
    
    # Tampilkan tabel (tanpa Branch dan Hari Libur)
    display_cols = [
        'Employee ID', 'Full Name', 'Organization', 'Job Position',
        'Work Days Bulan Ini', 'Jumlah Hadir', 'Jumlah Absen', 'Jumlah Cuti',
        'Jumlah Late In', 'Jumlah Early Out',
        'Total Jam Kerja (Real) Formatted', 'Total Jam Kerja (Plant) Formatted',
        'Checklist Plant', 'Kekurangan Jam Kerja Formatted',
        'Total Jam Late In Formatted', 'Total Jam Early Out Formatted'
    ]
    
    # Rename untuk display
    display_df = employee_stats[display_cols].copy()
    display_df.columns = [
        'ID', 'Nama', 'Organization', 'Posisi',
        'Work Days Bulan Ini', 'Jumlah Hadir', 'Jumlah Absen', 'Cuti',
        'Late In', 'Early Out',
        'Total Jam Kerja (Real)', 'Total Jam Kerja (Plant)',
        'Checklist Plant', 'Kekurangan Jam Kerja',
        'Total Jam Late In', 'Total Jam Early Out'
    ]
    
    # Search box
    search_term = st.text_input("🔍 Cari Karyawan (Nama atau ID)", "")
    if search_term:
        mask = (
            display_df['Nama'].str.contains(search_term, case=False, na=False) |
            display_df['ID'].astype(str).str.contains(search_term, case=False, na=False)
        )
        display_df = display_df[mask]
    
    st.dataframe(
        display_df,
        use_container_width=True,
        height=400
    )
    
    # Download button (hapus kolom Branch dan Organization dari CSV)
    csv_df = employee_stats.drop(columns=['Branch', 'Organization'], errors='ignore')
    csv = csv_df.to_csv(index=False)
    
    col_analysis1, col_analysis2 = st.columns(2)
    with col_analysis1:
        st.download_button(
            label="📥 Download Data Analisis (CSV)",
            data=csv,
            file_name=f"analisis_absensi_januari_{selected_branch}_{selected_org}.csv",
            mime="text/csv",
            key='download_analysis_csv'
        )
    with col_analysis2:
        # Siapkan data untuk PDF (hapus kolom Branch dan Organization)
        pdf_analysis_df = display_df.drop(columns=['Branch', 'Organization'], errors='ignore').copy()
        pdf_analysis = create_table_pdf(
            pdf_analysis_df,
            "ANALISIS PER KARYAWAN",
            f"Branch: {selected_branch} | Organization: {selected_org if selected_org != 'All' else 'Semua'}"
        )
        st.download_button(
            label="📄 Download Data Analisis (PDF)",
            data=pdf_analysis.getvalue(),
            file_name=f"analisis_absensi_januari_{selected_branch}_{selected_org}.pdf",
            mime="application/pdf",
            key='download_analysis_pdf'
        )
    
    st.markdown("---")
    
    # Tabel Checklist Compliance
    st.header("✅ Tabel Checklist Compliance")
    st.markdown("Checklist untuk memverifikasi compliance karyawan terhadap standar kerja")
    
    # Buat tabel checklist
    checklist_data = filtered_df.copy()
    
    # Filter hanya yang hadir
    checklist_data = checklist_data[checklist_data['Is Present'] == True].copy()
    
    # Fungsi untuk convert waktu ke menit (untuk perbandingan)
    def time_to_minutes(time_str):
        """Convert waktu HH:MM ke total menit"""
        if pd.isna(time_str) or time_str == '' or str(time_str).strip() == '':
            return None
        try:
            time_clean = str(time_str).strip()
            parts = time_clean.split(':')
            if len(parts) == 2:
                hour = int(parts[0])
                minute = int(parts[1])
                return hour * 60 + minute
            return None
        except:
            return None
    
    # Checklist 1: Kerja selama 8 jam per hari
    checklist_data['Checklist_8_Jam'] = checklist_data['Real Working Hour Decimal'].apply(
        lambda x: '✅' if x >= 8.0 else '❌'
    )
    
    # Checklist 2: Masuk <= 08:00 dan pulang >= 17:00
    # Logic: Masuk kurang dari atau sama dengan 08:00 = OK, lebih dari 08:00 = Tidak OK
    #        Pulang lebih dari atau sama dengan 17:00 = OK, kurang dari 17:00 = Tidak OK
    def check_in_out_time(row):
        check_in_minutes = time_to_minutes(row['Check In'])
        check_out_minutes = time_to_minutes(row['Check Out'])
        
        if check_in_minutes is None or check_out_minutes is None:
            return '❌'
        
        # Target: Check In <= 08:00 (480 menit) dan Check Out >= 17:00 (1020 menit)
        target_check_in = 8 * 60 + 0  # 08:00 = 480 menit
        target_check_out = 17 * 60 + 0  # 17:00 = 1020 menit
        
        if check_in_minutes <= target_check_in and check_out_minutes >= target_check_out:
            return '✅'
        else:
            return '❌'
    
    checklist_data['Checklist_Jam_8_17'] = checklist_data.apply(check_in_out_time, axis=1)
    
    # Pilih kolom untuk checklist (dengan Branch dan Organization untuk display di UI)
    checklist_display = checklist_data[[
        'Date', 'Employee ID', 'Full Name', 'Branch', 'Organization', 'Job Position',
        'Shift', 'Check In', 'Check Out', 'Real Working Hour', 'Real Working Hour Decimal',
        'Checklist_8_Jam', 'Checklist_Jam_8_17'
    ]].copy()
    
    # Filter tanggal untuk checklist
    col_check1, col_check2 = st.columns(2)
    with col_check1:
        min_date_check = checklist_display['Date'].min().date() if not checklist_display['Date'].empty else pd.Timestamp.now().date()
        max_date_check = checklist_display['Date'].max().date() if not checklist_display['Date'].empty else pd.Timestamp.now().date()
        date_start_check = st.date_input("Tanggal Mulai - Checklist", value=min_date_check, min_value=min_date_check, max_value=max_date_check, key='check_start')
    with col_check2:
        date_end_check = st.date_input("Tanggal Akhir - Checklist", value=max_date_check, min_value=min_date_check, max_value=max_date_check, key='check_end')
    
    # Filter berdasarkan tanggal
    checklist_display_filtered = checklist_display[
        (checklist_display['Date'].dt.date >= date_start_check) &
        (checklist_display['Date'].dt.date <= date_end_check)
    ].copy()
    
    # Search box untuk checklist (sebelum rename)
    search_checklist = st.text_input("🔍 Cari Karyawan (Nama atau ID) - Checklist", "", key='search_checklist')
    if search_checklist:
        mask_check = (
            checklist_display_filtered['Full Name'].str.contains(search_checklist, case=False, na=False) |
            checklist_display_filtered['Employee ID'].astype(str).str.contains(search_checklist, case=False, na=False)
        )
        checklist_display_filtered = checklist_display_filtered[mask_check]
    
    # Sort berdasarkan nama dulu, baru tanggal (untuk grouping per orang)
    # Simpan Date as datetime dulu untuk sorting yang benar
    checklist_display_filtered = checklist_display_filtered.sort_values(['Full Name', 'Date'], ascending=[True, True])
    
    # Format tanggal untuk display (hanya tanggal, tanpa jam)
    checklist_display_filtered['Date'] = checklist_display_filtered['Date'].dt.strftime('%Y-%m-%d')
    # Pastikan kolom Date adalah string, bukan datetime
    checklist_display_filtered['Date'] = checklist_display_filtered['Date'].astype(str)
    
    # Rename kolom untuk display
    checklist_display_filtered.columns = [
        'Tanggal', 'ID', 'Nama', 'Branch', 'Organization', 'Posisi',
        'Shift', 'Check In', 'Check Out', 'Jam Kerja (Format)', 'Jam Kerja (Desimal)',
        '✅ Kerja 8 Jam/Hari', '✅ Masuk 08:00 & Pulang 17:00'
    ]
    
    # Buat copy untuk download (tanpa Branch dan Organization)
    # Pastikan sudah di-sort per orang (Nama, lalu Tanggal)
    checklist_display_for_download = checklist_display_filtered.drop(columns=['Branch', 'Organization'], errors='ignore').copy()
    # Sort ulang untuk download: Nama dulu, baru Tanggal (ascending)
    checklist_display_for_download = checklist_display_for_download.sort_values(['Nama', 'Tanggal'], ascending=[True, True])
    
    # Statistik checklist
    total_checklist = len(checklist_display_filtered)
    compliant_8jam = len(checklist_display_filtered[checklist_display_filtered['✅ Kerja 8 Jam/Hari'] == '✅'])
    compliant_8_17 = len(checklist_display_filtered[checklist_display_filtered['✅ Masuk 08:00 & Pulang 17:00'] == '✅'])
    compliant_both = len(checklist_display_filtered[
        (checklist_display_filtered['✅ Kerja 8 Jam/Hari'] == '✅') &
        (checklist_display_filtered['✅ Masuk 08:00 & Pulang 17:00'] == '✅')
    ])
    
    col_check_stat1, col_check_stat2, col_check_stat3, col_check_stat4 = st.columns(4)
    with col_check_stat1:
        st.metric("Total Record", total_checklist)
    with col_check_stat2:
        st.metric("✅ Kerja 8 Jam/Hari", f"{compliant_8jam} ({(compliant_8jam/total_checklist*100) if total_checklist > 0 else 0:.1f}%)")
    with col_check_stat3:
        st.metric("✅ Masuk 08:00 & Pulang 17:00", f"{compliant_8_17} ({(compliant_8_17/total_checklist*100) if total_checklist > 0 else 0:.1f}%)")
    with col_check_stat4:
        st.metric("✅ Keduanya Compliant", f"{compliant_both} ({(compliant_both/total_checklist*100) if total_checklist > 0 else 0:.1f}%)")
    
    # Tampilkan tabel checklist
    st.dataframe(
        checklist_display_filtered,
        use_container_width=True,
        height=500
    )
    
    # Download button untuk checklist (tanpa Branch dan Organization)
    csv_checklist = checklist_display_for_download.to_csv(index=False)
    
    col_checklist1, col_checklist2 = st.columns(2)
    with col_checklist1:
        st.download_button(
            label="📥 Download Tabel Checklist (CSV)",
            data=csv_checklist,
            file_name=f"checklist_compliance_ho_jakarta_{date_start_check}_{date_end_check}.csv",
            mime="text/csv",
            key='download_checklist_csv'
        )
    with col_checklist2:
        pdf_checklist = create_table_pdf(
            checklist_display_for_download,
            "CHECKLIST COMPLIANCE",
            f"Branch: {selected_branch} | Periode: {date_start_check} - {date_end_check}"
        )
        st.download_button(
            label="📄 Download Tabel Checklist (PDF)",
            data=pdf_checklist.getvalue(),
            file_name=f"checklist_compliance_ho_jakarta_{date_start_check}_{date_end_check}.pdf",
            mime="application/pdf",
            key='download_checklist_pdf'
        )
    
    st.markdown("---")
    
    # Visualisasi
    st.header("📊 Visualisasi Data")
    
    # Tab untuk berbagai visualisasi
    tab1, tab2, tab3, tab4 = st.tabs([
        "📈 Jumlah Absensi",
        "⏰ Keterlambatan",
        "🚪 Early Out",
        "⏱️ Jam Kerja"
    ])
    
    with tab1:
        st.subheader("Jumlah Hadir Per Karyawan")
        # Top 20 karyawan dengan hadir terbanyak
        top_attendance = employee_stats.nlargest(20, 'Jumlah Hadir')
        fig1 = px.bar(
            top_attendance,
            x='Jumlah Hadir',
            y='Full Name',
            orientation='h',
            title="Top 20 Karyawan dengan Kehadiran Terbanyak",
            labels={'Full Name': 'Nama Karyawan', 'Jumlah Hadir': 'Jumlah Hadir'}
        )
        fig1.update_layout(height=600, yaxis={'categoryorder': 'total ascending'})
        st.plotly_chart(fig1, use_container_width=True)
        
        # Download data untuk chart ini
        csv_viz_attendance_tab = top_attendance[['Employee ID', 'Full Name', 'Jumlah Hadir']].to_csv(index=False)
        st.download_button(
            label="📥 Download Data Chart Ini (CSV)",
            data=csv_viz_attendance_tab,
            file_name=f"data_chart_jumlah_hadir_tab_{selected_branch}.csv",
            mime="text/csv",
            key='download_viz_attendance_tab1'
        )
    
    with tab2:
        st.subheader("Keterlambatan Per Karyawan")
        # Top 20 karyawan dengan late in terbanyak
        top_late = employee_stats[employee_stats['Jumlah Late In'] > 0].nlargest(20, 'Jumlah Late In')
        if len(top_late) > 0:
            fig2 = px.bar(
                top_late,
                x='Jumlah Late In',
                y='Full Name',
                orientation='h',
                title="Top 20 Karyawan dengan Keterlambatan Terbanyak",
                labels={'Full Name': 'Nama Karyawan', 'Jumlah Late In': 'Jumlah Keterlambatan'},
                color='Jumlah Late In',
                color_continuous_scale='Reds'
            )
            fig2.update_layout(height=600, yaxis={'categoryorder': 'total ascending'})
            st.plotly_chart(fig2, use_container_width=True)
            
            # Download data untuk chart ini
            csv_viz_late_tab = top_late[['Employee ID', 'Full Name', 'Jumlah Late In']].to_csv(index=False)
            st.download_button(
                label="📥 Download Data Chart Ini (CSV)",
                data=csv_viz_late_tab,
                file_name=f"data_chart_late_in_tab_{selected_branch}.csv",
                mime="text/csv",
                key='download_viz_late_tab2'
            )
        else:
            st.info("Tidak ada data keterlambatan")
    
    with tab3:
        st.subheader("Early Out Per Karyawan")
        # Top 20 karyawan dengan early out terbanyak
        top_early = employee_stats[employee_stats['Jumlah Early Out'] > 0].nlargest(20, 'Jumlah Early Out')
        if len(top_early) > 0:
            fig3 = px.bar(
                top_early,
                x='Jumlah Early Out',
                y='Full Name',
                orientation='h',
                title="Top 20 Karyawan dengan Early Out Terbanyak",
                labels={'Full Name': 'Nama Karyawan', 'Jumlah Early Out': 'Jumlah Early Out'},
                color='Jumlah Early Out',
                color_continuous_scale='Oranges'
            )
            fig3.update_layout(height=600, yaxis={'categoryorder': 'total ascending'})
            st.plotly_chart(fig3, use_container_width=True)
        else:
            st.info("Tidak ada data early out")
        
        # Download data Early Out
        if len(top_early) > 0:
            csv_viz_early = top_early[['Employee ID', 'Full Name', 'Jumlah Early Out']].to_csv(index=False)
            st.download_button(
                label="📥 Download Data Chart Early Out (CSV)",
                data=csv_viz_early,
                file_name=f"data_chart_early_out_{selected_branch}.csv",
                mime="text/csv",
                key='download_viz_early'
            )
    
    with tab4:
        st.subheader("Total Jam Kerja Per Karyawan")
        # Top 20 karyawan dengan jam kerja terbanyak
        top_hours = employee_stats.nlargest(20, 'Total Jam Kerja (Real)')
        fig4 = px.bar(
            top_hours,
            x='Total Jam Kerja (Real)',
            y='Full Name',
            orientation='h',
            title="Top 20 Karyawan dengan Jam Kerja Terbanyak",
            labels={'Full Name': 'Nama Karyawan', 'Total Jam Kerja (Real)': 'Total Jam Kerja'},
            color='Total Jam Kerja (Real)',
            color_continuous_scale='Greens'
        )
        fig4.update_layout(height=600, yaxis={'categoryorder': 'total ascending'})
        st.plotly_chart(fig4, use_container_width=True)
        
        # Download data untuk chart jam kerja
        csv_viz_hours = top_hours[['Employee ID', 'Full Name', 'Total Jam Kerja (Real)', 'Total Jam Kerja (Real) Formatted']].to_csv(index=False)
        st.download_button(
            label="📥 Download Data Chart Jam Kerja (CSV)",
            data=csv_viz_hours,
            file_name=f"data_chart_jam_kerja_{selected_branch}.csv",
            mime="text/csv",
            key='download_viz_hours'
        )
        
        # Distribusi jam kerja
        st.subheader("Distribusi Total Jam Kerja")
        fig5 = px.histogram(
            employee_stats,
            x='Total Jam Kerja (Real)',
            nbins=30,
            title="Distribusi Total Jam Kerja Semua Karyawan",
            labels={'Total Jam Kerja (Real)': 'Total Jam Kerja', 'count': 'Jumlah Karyawan'}
        )
        st.plotly_chart(fig5, use_container_width=True)
        
        # Download data distribusi jam kerja
        distribution_data = employee_stats[['Employee ID', 'Full Name', 'Total Jam Kerja (Real)', 'Total Jam Kerja (Real) Formatted']].copy()
        csv_distribution = distribution_data.to_csv(index=False)
        st.download_button(
            label="📥 Download Data Distribusi Jam Kerja (CSV)",
            data=csv_distribution,
            file_name=f"data_distribusi_jam_kerja_{selected_branch}.csv",
            mime="text/csv",
            key='download_distribution'
        )
    
    st.markdown("---")
    
    # Detail per karyawan
    st.header("🔍 Detail Per Karyawan")
    
    # Search box untuk memilih karyawan
    employee_list = sorted(employee_stats['Full Name'].unique())
    search_employee = st.text_input("🔍 Cari Karyawan", "", placeholder="Ketik nama atau ID karyawan...")
    
    if search_employee:
        filtered_employees = [
            emp for emp in employee_list 
            if search_employee.lower() in emp.lower() or 
            search_employee in str(employee_stats[employee_stats['Full Name'] == emp]['Employee ID'].iloc[0])
        ]
        if filtered_employees:
            selected_employee = st.selectbox(
                "Pilih Karyawan",
                options=filtered_employees,
                key='employee_select'
            )
        else:
            st.warning("Karyawan tidak ditemukan")
            selected_employee = None
    else:
        selected_employee = st.selectbox(
            "Pilih Karyawan",
            options=employee_list,
            key='employee_select'
        )
    
    if selected_employee:
        emp_data = employee_stats[employee_stats['Full Name'] == selected_employee].iloc[0]
        emp_detail = filtered_df[filtered_df['Full Name'] == selected_employee].copy()
        
        # Header dengan informasi karyawan
        st.markdown("### 👤 Informasi Karyawan")
        info_col1, info_col2, info_col3, info_col4 = st.columns(4)
        
        with info_col1:
            st.markdown(f"**🆔 Employee ID**  \n{emp_data['Employee ID']}")
        with info_col2:
            st.markdown(f"**🏢 Branch**  \n{emp_data['Branch']}")
        with info_col3:
            st.markdown(f"**🏛️ Organization**  \n{emp_data['Organization']}")
        with info_col4:
            st.markdown(f"**💼 Posisi**  \n{emp_data['Job Position']}")
        
        st.markdown("---")
        
        # Statistik utama dengan cards yang lebih menarik
        st.markdown("### 📊 Statistik Absensi Bulan Ini")
        
        # Row 1: Kehadiran
        stat_col1, stat_col2, stat_col3, stat_col4 = st.columns(4)
        with stat_col1:
            attendance_rate = (emp_data['Jumlah Hadir'] / emp_data['Work Days Bulan Ini'] * 100) if emp_data['Work Days Bulan Ini'] > 0 else 0
            st.metric(
                "📅 Work Days Bulan Ini",
                int(emp_data['Work Days Bulan Ini']),
                help="Hari kerja yang seharusnya"
            )
        with stat_col2:
            st.metric(
                "✅ Jumlah Hadir",
                int(emp_data['Jumlah Hadir']),
                delta=f"{attendance_rate:.1f}%",
                delta_color="normal",
                help="Total hari hadir"
            )
        with stat_col3:
            st.metric(
                "❌ Jumlah Absen",
                int(emp_data['Jumlah Absen']),
                delta_color="inverse",
                help="Total hari tidak hadir"
            )
        with stat_col4:
            st.metric(
                "🏖️ Hari Libur",
                int(emp_data['Jumlah Hari Libur']),
                help="Total hari libur"
            )
        
        # Row 2: Cuti dan Keterlambatan
        stat_col5, stat_col6, stat_col7, stat_col8 = st.columns(4)
        with stat_col5:
            st.metric(
                "✈️ Cuti",
                int(emp_data['Jumlah Cuti']),
                help="Total hari cuti"
            )
        with stat_col6:
            st.metric(
                "⏰ Late In",
                int(emp_data['Jumlah Late In']),
                delta_color="inverse",
                help="Jumlah keterlambatan"
            )
        with stat_col7:
            st.metric(
                "🚪 Early Out",
                int(emp_data['Jumlah Early Out']),
                delta_color="inverse",
                help="Jumlah pulang lebih cepat"
            )
        with stat_col8:
            st.metric(
                "⏱️ Total Jam Kerja",
                emp_data['Total Jam Kerja (Real) Formatted'],
                help="Total jam kerja (Real Working Hour)"
            )
        
        st.markdown("---")
        
        # Visualisasi ringkasan
        viz_col1, viz_col2 = st.columns(2)
        
        with viz_col1:
            # Pie chart untuk distribusi kehadiran
            attendance_data = {
                'Hadir': int(emp_data['Jumlah Hadir']),
                'Absen': int(emp_data['Jumlah Absen']),
                'Cuti': int(emp_data['Jumlah Cuti']),
                'Hari Libur': int(emp_data['Jumlah Hari Libur'])
            }
            fig_pie = px.pie(
                values=list(attendance_data.values()),
                names=list(attendance_data.keys()),
                title="Distribusi Kehadiran Bulan Ini",
                color_discrete_map={
                    'Hadir': '#2ecc71',
                    'Absen': '#e74c3c',
                    'Cuti': '#3498db',
                    'Hari Libur': '#f39c12'
                }
            )
            fig_pie.update_layout(height=350)
            st.plotly_chart(fig_pie, use_container_width=True)
        
        with viz_col2:
            # Bar chart untuk late in dan early out
            violation_data = {
                'Late In': int(emp_data['Jumlah Late In']),
                'Early Out': int(emp_data['Jumlah Early Out'])
            }
            fig_bar = px.bar(
                x=list(violation_data.keys()),
                y=list(violation_data.values()),
                title="Pelanggaran Waktu Kerja",
                labels={'x': 'Jenis Pelanggaran', 'y': 'Jumlah'},
                color=list(violation_data.keys()),
                color_discrete_map={
                    'Late In': '#e74c3c',
                    'Early Out': '#f39c12'
                }
            )
            fig_bar.update_layout(height=350, showlegend=False)
            st.plotly_chart(fig_bar, use_container_width=True)
        
        st.markdown("---")
        
        # Tabel detail harian dengan filter dan format yang lebih baik
        st.markdown("### 📋 Detail Harian")
        
        # Filter tanggal untuk detail harian
        detail_date_col1, detail_date_col2 = st.columns(2)
        with detail_date_col1:
            min_date_detail = emp_detail['Date'].min().date() if not emp_detail['Date'].empty else pd.Timestamp.now().date()
            max_date_detail = emp_detail['Date'].max().date() if not emp_detail['Date'].empty else pd.Timestamp.now().date()
            detail_date_start = st.date_input("Tanggal Mulai", value=min_date_detail, min_value=min_date_detail, max_value=max_date_detail, key='detail_start')
        with detail_date_col2:
            detail_date_end = st.date_input("Tanggal Akhir", value=max_date_detail, min_value=min_date_detail, max_value=max_date_detail, key='detail_end')
        
        # Filter berdasarkan tanggal
        emp_detail_filtered = emp_detail[
            (emp_detail['Date'].dt.date >= detail_date_start) &
            (emp_detail['Date'].dt.date <= detail_date_end)
        ].copy()
        
        # Tambahkan status untuk setiap hari
        def get_status(row):
            if row['Is Present']:
                return '✅ Hadir'
            elif row['Is Leave']:
                return '✈️ Cuti'
            elif row['Is Dayoff']:
                return '🏖️ Hari Libur'
            elif row['Is Absent']:
                return '❌ Absen'
            else:
                return '❓ Tidak Diketahui'
        
        emp_detail_filtered['Status'] = emp_detail_filtered.apply(get_status, axis=1)
        
        # Pilih kolom untuk ditampilkan
        detail_cols = ['Date', 'Status', 'Shift', 'Check In', 'Check Out', 'Late In', 'Early Out', 
                      'Real Working Hour', 'Actual Working Hour', 'Attendance Code']
        detail_display = emp_detail_filtered[detail_cols].copy()
        detail_display = detail_display.sort_values('Date', ascending=False)
        
        # Format tanggal
        detail_display['Date'] = detail_display['Date'].dt.strftime('%Y-%m-%d (%A)')
        
        # Rename kolom
        detail_display.columns = [
            'Tanggal', 'Status', 'Shift', 'Check In', 'Check Out', 'Late In', 'Early Out',
            'Jam Kerja (Real)', 'Jam Kerja (Actual)', 'Kode Absensi'
        ]
        
        # Filter berdasarkan status
        status_filter = st.multiselect(
            "Filter Status",
            options=['✅ Hadir', '✈️ Cuti', '🏖️ Hari Libur', '❌ Absen'],
            default=['✅ Hadir', '✈️ Cuti', '🏖️ Hari Libur', '❌ Absen'],
            key='status_filter'
        )
        
        if status_filter:
            detail_display = detail_display[detail_display['Status'].isin(status_filter)]
        
        # Tampilkan tabel
        st.dataframe(
            detail_display,
            use_container_width=True,
            height=400,
            hide_index=True
        )
        
        # Download button untuk detail harian
        csv_detail = detail_display.to_csv(index=False)
        
        col_detail1, col_detail2 = st.columns(2)
        with col_detail1:
            st.download_button(
                label="📥 Download Detail Harian (CSV)",
                data=csv_detail,
                file_name=f"detail_harian_{selected_employee.replace(' ', '_')}_{detail_date_start}_{detail_date_end}.csv",
                mime="text/csv",
                key='download_detail_csv'
            )
        with col_detail2:
            pdf_detail = create_table_pdf(
                detail_display,
                f"DETAIL HARIAN - {selected_employee}",
                f"Periode: {detail_date_start} - {detail_date_end}"
            )
            st.download_button(
                label="📄 Download Detail Harian (PDF)",
                data=pdf_detail.getvalue(),
                file_name=f"detail_harian_{selected_employee.replace(' ', '_')}_{detail_date_start}_{detail_date_end}.pdf",
                mime="application/pdf",
                key='download_detail_pdf'
            )
        
        st.markdown("---")
        
        # Download button untuk seluruh section Detail Per Karyawan
        st.markdown("### 📥 Download Report Lengkap Detail Karyawan")
        st.markdown("Download seluruh informasi karyawan termasuk informasi, statistik, dan detail harian")
        
        # Siapkan data lengkap untuk download
        # 1. Informasi Karyawan
        employee_info = pd.DataFrame({
            'Informasi': ['Employee ID', 'Nama', 'Branch', 'Organization', 'Posisi'],
            'Nilai': [
                emp_data['Employee ID'],
                emp_data['Full Name'],
                emp_data['Branch'],
                emp_data['Organization'],
                emp_data['Job Position']
            ]
        })
        
        # 2. Statistik Absensi
        attendance_stats = pd.DataFrame({
            'Metrik': [
                'Work Days Bulan Ini', 'Jumlah Hadir', 'Jumlah Absen', 'Hari Libur', 'Cuti',
                'Jumlah Late In', 'Jumlah Early Out', 'Total Jam Kerja (Real)', 'Total Jam Kerja (Actual)'
            ],
            'Nilai': [
                int(emp_data['Work Days Bulan Ini']),
                int(emp_data['Jumlah Hadir']),
                int(emp_data['Jumlah Absen']),
                int(emp_data['Jumlah Hari Libur']),
                int(emp_data['Jumlah Cuti']),
                int(emp_data['Jumlah Late In']),
                int(emp_data['Jumlah Early Out']),
                emp_data['Total Jam Kerja (Real) Formatted'],
                emp_data['Total Jam Kerja (Actual) Formatted']
            ]
        })
        
        # Download buttons untuk report lengkap
        col_full1, col_full2 = st.columns(2)
        with col_full1:
            # CSV - gabungkan semua data
            csv_combined_rows = []
            
            # Tambahkan informasi karyawan
            for _, row in employee_info.iterrows():
                csv_combined_rows.append({
                    'Section': 'INFORMASI KARYAWAN',
                    'Kolom': row['Informasi'],
                    'Nilai': row['Nilai']
                })
            
            # Tambahkan statistik absensi
            for _, row in attendance_stats.iterrows():
                csv_combined_rows.append({
                    'Section': 'STATISTIK ABSENSI',
                    'Kolom': row['Metrik'],
                    'Nilai': row['Nilai']
                })
            
            # Tambahkan detail harian
            detail_for_csv = detail_display.copy()
            detail_for_csv.insert(0, 'Section', 'DETAIL HARIAN')
            for _, row in detail_for_csv.iterrows():
                row_dict = {'Section': 'DETAIL HARIAN'}
                for col in detail_for_csv.columns:
                    if col != 'Section':
                        row_dict[col] = row[col]
                csv_combined_rows.append(row_dict)
            
            csv_combined = pd.DataFrame(csv_combined_rows)
            csv_full = csv_combined.to_csv(index=False)
            st.download_button(
                label="📥 Download Report Lengkap (CSV)",
                data=csv_full,
                file_name=f"report_lengkap_{selected_employee.replace(' ', '_')}_{detail_date_start}_{detail_date_end}.csv",
                mime="text/csv",
                key='download_full_report_csv'
            )
        
        with col_full2:
            # PDF - buat PDF lengkap dengan semua informasi
            pdf_full = create_employee_full_pdf(
                employee_info,
                attendance_stats,
                detail_display,
                selected_employee,
                emp_data,
                detail_date_start,
                detail_date_end
            )
            st.download_button(
                label="📄 Download Report Lengkap (PDF)",
                data=pdf_full.getvalue(),
                file_name=f"report_lengkap_{selected_employee.replace(' ', '_')}_{detail_date_start}_{detail_date_end}.pdf",
                mime="application/pdf",
                key='download_full_report_pdf'
            )
    
    # Tombol Download Report Lengkap di Sidebar (setelah semua data diproses)
    st.sidebar.markdown("---")
    st.sidebar.header("📥 Download Report")
    
    # Siapkan data untuk report
    checklist_report_data = filtered_df[filtered_df['Is Present'] == True].copy()
    
    # Fungsi untuk convert waktu ke menit
    def time_to_minutes_report(time_str):
        if pd.isna(time_str) or time_str == '' or str(time_str).strip() == '':
            return None
        try:
            time_clean = str(time_str).strip()
            parts = time_clean.split(':')
            if len(parts) == 2:
                hour = int(parts[0])
                minute = int(parts[1])
                return hour * 60 + minute
            return None
        except:
            return None
    
    # Checklist 1: Kerja 8 jam
    if len(checklist_report_data) > 0:
        checklist_report_data['Checklist_8_Jam'] = checklist_report_data['Real Working Hour Decimal'].apply(
            lambda x: '✅' if x >= 8.0 else '❌'
        )
        
        # Checklist 2: Masuk <= 08:00 dan pulang >= 17:00
        def check_in_out_time_report(row):
            check_in_minutes = time_to_minutes_report(row['Check In'])
            check_out_minutes = time_to_minutes_report(row['Check Out'])
            
            if check_in_minutes is None or check_out_minutes is None:
                return '❌'
            
            target_check_in = 8 * 60 + 0
            target_check_out = 17 * 60 + 0
            
            if check_in_minutes <= target_check_in and check_out_minutes >= target_check_out:
                return '✅'
            else:
                return '❌'
        
        checklist_report_data['Checklist_Jam_8_17'] = checklist_report_data.apply(check_in_out_time_report, axis=1)
        
        # Siapkan checklist untuk report (tanpa Branch dan Organization)
        checklist_report = checklist_report_data[[
            'Date', 'Employee ID', 'Full Name', 'Job Position',
            'Shift', 'Check In', 'Check Out', 'Real Working Hour', 'Real Working Hour Decimal',
            'Checklist_8_Jam', 'Checklist_Jam_8_17'
        ]].copy()
        # Sort berdasarkan nama dulu, baru tanggal (untuk grouping per orang)
        checklist_report = checklist_report.sort_values(['Full Name', 'Date'], ascending=[True, True])
        # Format tanggal menjadi string (hanya tanggal, tanpa jam)
        checklist_report['Date'] = checklist_report['Date'].dt.strftime('%Y-%m-%d')
        # Pastikan kolom Date adalah string, bukan datetime
        checklist_report['Date'] = checklist_report['Date'].astype(str)
        checklist_report.columns = [
            'Tanggal', 'ID', 'Nama', 'Posisi',
            'Shift', 'Check In', 'Check Out', 'Jam Kerja (Format)', 'Jam Kerja (Desimal)',
            '✅ Kerja 8 Jam/Hari', '✅ Masuk 08:00 & Pulang 17:00'
        ]
    else:
        checklist_report = pd.DataFrame()
    
    # Buat report Excel
    excel_file = create_excel_report(employee_stats, checklist_report, filtered_df, selected_branch, selected_org)
    
    # Buat report PDF
    pdf_file = create_pdf_report(employee_stats, checklist_report, filtered_df, selected_branch, selected_org)
    
    # Download buttons
    org_name = selected_org if selected_org != 'All' else 'All'
    file_name = f"Report_Absensi_{selected_branch}_{org_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    pdf_file_name = f"Report_Absensi_{selected_branch}_{org_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    st.sidebar.download_button(
        label="📊 Download Report Excel",
        data=excel_file.getvalue(),
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        key='download_excel'
    )
    
    st.sidebar.download_button(
        label="📄 Download Report PDF",
        data=pdf_file.getvalue(),
        file_name=pdf_file_name,
        mime="application/pdf",
        use_container_width=True,
        key='download_pdf'
    )
    
    st.sidebar.markdown("**Report berisi:**")
    st.sidebar.markdown("""
    - 📊 Ringkasan Statistik
    - 👥 Analisis Per Karyawan
    - ✅ Checklist Compliance
    - 📋 Detail Harian (Excel only)
    """)

else:
    st.error("Gagal memuat data. Pastikan file january.csv ada di direktori yang sama.")

