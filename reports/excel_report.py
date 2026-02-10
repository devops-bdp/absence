"""Excel report generator"""
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from utils.formatters import format_hours_excel


def create_excel_report(employee_stats_df, checklist_df, filtered_df, branch, org):
    """Membuat report Excel lengkap dengan beberapa sheet"""
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    
    # Sheet 1: Ringkasan Statistik
    total_emp = filtered_df['Employee ID'].nunique()
    total_pres = filtered_df['Is Present'].sum()
    total_abs = filtered_df['Is Absent'].sum()
    work_days = employee_stats_df['Work Days Bulan Ini'].iloc[0] if len(employee_stats_df) > 0 else 0
    total_work_day = total_emp * work_days
    attendance_pct = (total_pres / total_work_day * 100) if total_work_day > 0 else 0
    absent_pct = (total_abs / total_work_day * 100) if total_work_day > 0 else 0
    
    # Hitung total jam kerja
    total_jam_kerja_real = employee_stats_df['Total Jam Kerja (Real)'].sum() if 'Total Jam Kerja (Real)' in employee_stats_df.columns else filtered_df['Real Working Hour Decimal'].sum()
    total_jam_kerja_plant = total_work_day * 8
    total_jam_kerja_real_formatted = format_hours_excel(total_jam_kerja_real)
    total_jam_kerja_plant_formatted = format_hours_excel(total_jam_kerja_plant)
    
    summary_data = {
        'Metrik': [
            'Total Karyawan', 'Work Day', 'Total Work Day', 'Total Kehadiran',
            'Total Kehadiran (%)', 'Total Tidak Hadir', 'Total Tidak Hadir (%)',
            'Total Jam Kerja', 'Plan Jam Kerja'
        ],
        'Nilai': [
            total_emp, work_days, total_work_day, total_pres,
            f"{attendance_pct:.2f}%", total_abs, f"{absent_pct:.2f}%",
            total_jam_kerja_real_formatted, total_jam_kerja_plant_formatted
        ]
    }
    summary_df = pd.DataFrame(summary_data)
    summary_df.to_excel(writer, sheet_name='Ringkasan Statistik', index=False)
    
    # Sheet 2: Analisis Per Karyawan
    employee_stats_df_export = employee_stats_df.drop(columns=['Branch', 'Organization'], errors='ignore')
    employee_stats_df_export.to_excel(writer, sheet_name='Analisis Per Karyawan', index=False)
    
    # Sheet 3: Checklist Compliance
    checklist_df_export = checklist_df.drop(columns=['Branch', 'Organization'], errors='ignore')
    checklist_df_export.to_excel(writer, sheet_name='Checklist Compliance', index=False)
    
    # Sheet 4: Detail Harian
    detail_cols = ['Date', 'Employee ID', 'Full Name', 'Job Position',
                  'Shift', 'Check In', 'Check Out', 'Late In', 'Early Out',
                  'Real Working Hour', 'Actual Working Hour', 'Attendance Code',
                  'Is Present', 'Is Absent', 'Is Dayoff', 'Is Leave']
    detail_df = filtered_df[detail_cols].copy()
    detail_df = detail_df.sort_values(['Full Name', 'Date'], ascending=[True, False])
    detail_df['Date'] = detail_df['Date'].dt.strftime('%Y-%m-%d')
    detail_df.to_excel(writer, sheet_name='Detail Harian', index=False)
    
    # Formatting Excel
    workbook = writer.book
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Format Sheet Ringkasan Statistik
    ws1 = workbook['Ringkasan Statistik']
    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 15
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Format Sheet Analisis Per Karyawan
    ws2 = workbook['Analisis Per Karyawan']
    for col in range(1, ws2.max_column + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 18
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Format Sheet Checklist Compliance
    ws3 = workbook['Checklist Compliance']
    col_widths_checklist = {
        'Tanggal': 12, 'ID': 10, 'Nama': 25, 'Posisi': 20, 'Shift': 15,
        'Check In': 10, 'Check Out': 10, 'Jam Kerja (Format)': 15,
        'Jam Kerja (Desimal)': 12, '✅ Kerja 8 Jam/Hari': 18,
        '✅ Masuk 08:00 & Pulang 17:00': 25
    }
    for col_idx, col_name in enumerate(checklist_df_export.columns, start=1):
        col_letter = get_column_letter(col_idx)
        width = col_widths_checklist.get(col_name, 15)
        ws3.column_dimensions[col_letter].width = width
    for cell in ws3[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Format Sheet Detail Harian
    ws4 = workbook['Detail Harian']
    for col in range(1, ws4.max_column + 1):
        ws4.column_dimensions[get_column_letter(col)].width = 15
    for cell in ws4[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Freeze first row untuk semua sheet
    for ws in workbook.worksheets:
        ws.freeze_panes = 'A2'
    
    writer.close()
    output.seek(0)
    return output

